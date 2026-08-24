from __future__ import annotations

from pathlib import Path
import openpyxl
import pytest

from core.po_registry import FIXED_PO_DETAIL, FIXED_PO_SUB
from core.slip_printer_engine import START_ROW
from ui.app_controller import APP_TITLE, AppController
from ui.app_state import AppState


def _create_test_excel(file_path: Path, rows_data: list[dict[str, str]]) -> Path:
    """Create a valid Excel file matching the template schema starting at START_ROW."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for r_idx, data in enumerate(rows_data, start=START_ROW):
        ws.cell(row=r_idx, column=1, value=data.get("item_code", "ITEM001"))
        ws.cell(row=r_idx, column=2, value=data.get("item_name", "Test Item"))
        ws.cell(row=r_idx, column=3, value=data.get("total_qty", "100"))
        ws.cell(row=r_idx, column=4, value=data.get("carton_qty", "100"))
        ws.cell(row=r_idx, column=5, value=data.get("po", ""))
        ws.cell(row=r_idx, column=6, value=data.get("po_detail", "00010"))
        ws.cell(row=r_idx, column=7, value=data.get("po_sub", "+001"))
        ws.cell(row=r_idx, column=8, value=data.get("box", "001/001"))
        ws.cell(row=r_idx, column=9, value=data.get("rev", "01"))
        ws.cell(row=r_idx, column=10, value=data.get("lot", "LOT1234567"))

    wb.save(file_path)
    wb.close()
    return file_path


class MockView:
    def __init__(self, state: AppState):
        self.app_state = state
        self.logs: list[str] = []
        self.records_set: list = []

    def append_log(self, message: str) -> None:
        self.logs.append(message)

    def set_records(self, records, select_index=None, source=None) -> None:
        self.app_state.records = list(records)
        self.records_set = list(records)
        if source:
            self.append_log(source)


def test_import_from_excel_no_duplicates(tk_root, tmp_path, monkeypatch):
    """Test importing Excel when no duplicate EDI codes exist in po_registry."""
    excel_file = tmp_path / "no_dup.xlsx"
    _create_test_excel(
        excel_file,
        [
            {"item_code": "ITEM01", "item_name": "Part 1", "po": "PO_CLEAN_01", "po_detail": "00010", "po_sub": "+001", "box": "001/001"},
            {"item_code": "ITEM02", "item_name": "Part 2", "po": "PO_CLEAN_02", "po_detail": "00010", "po_sub": "+001", "box": "001/002"},
        ],
    )

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    # No duplicate warning should be triggered
    assert len(warnings) == 0
    assert len(state.records) == 2
    assert state.records[0].po == "PO_CLEAN_01"
    assert state.records[1].po == "PO_CLEAN_02"
    assert any("Đã import 2 dòng" in log for log in mock_view.logs)

    state.po_registry.close()


def test_import_from_excel_with_duplicates_shows_warning_and_loads_all(tk_root, tmp_path, monkeypatch):
    """Test importing Excel with existing duplicate EDI codes in po_registry:
    Shows non-blocking warning popup and loads all records into app_state.records.
    """
    excel_file = tmp_path / "with_dup.xlsx"
    _create_test_excel(
        excel_file,
        [
            {"item_code": "ITEM01", "item_name": "Part 1", "po": "PO_DUP_01", "po_detail": "00010", "po_sub": "+001", "box": "001/001"},
            {"item_code": "ITEM02", "item_name": "Part 2", "po": "PO_DUP_02", "po_detail": "00010", "po_sub": "+001", "box": "001/001"},
            {"item_code": "ITEM03", "item_name": "Part 3", "po": "PO_NEW_03", "po_detail": "00010", "po_sub": "+001", "box": "001/001"},
            {"item_code": "ITEM04", "item_name": "Part 4", "po": "PO_DUP_04", "po_detail": "00010", "po_sub": "+001", "box": "001/001"},
        ],
    )

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    # Pre-register 3 duplicate combos in po_registry
    state.po_registry.register_combo("PO_DUP_01", "00010", "+001", "001/001")
    state.po_registry.register_combo("PO_DUP_02", "00010", "+001", "001/001")
    state.po_registry.register_combo("PO_DUP_04", "00010", "+001", "001/001")

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    # 1. Verify warning was shown with correct title and duplicate count
    assert len(warnings) == 1
    title, msg = warnings[0]
    assert title == APP_TITLE
    assert "Phát hiện 3 dòng có mã EDI đã tồn tại" in msg
    assert "PO_DUP_01" in msg
    assert "PO_DUP_02" in msg
    assert "PO_DUP_04" in msg
    assert "dữ liệu vẫn được nạp vào bảng" in msg.lower()

    # 2. Verify all 4 records were loaded into state.records (non-blocking)
    assert len(state.records) == 4
    assert [r.po for r in state.records] == ["PO_DUP_01", "PO_DUP_02", "PO_NEW_03", "PO_DUP_04"]

    # 3. Verify logs captured the warning and successful import
    assert any("Cảnh báo: Phát hiện 3 dòng trùng mã EDI" in log for log in mock_view.logs)
    assert any("Đã import 4 dòng từ Excel" in log for log in mock_view.logs)

    state.po_registry.close()


def test_import_from_excel_more_than_three_duplicates(tk_root, tmp_path, monkeypatch):
    """Test importing Excel with >3 duplicates formats sample codes and truncation text properly."""
    rows = [
        {"item_code": f"ITEM{i:02d}", "item_name": f"Part {i}", "po": f"PO_MANY_{i}", "po_detail": "00010", "po_sub": "+001", "box": "001/001"}
        for i in range(1, 6)
    ]
    excel_file = tmp_path / "many_dups.xlsx"
    _create_test_excel(excel_file, rows)

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    # Pre-register all 5 combos
    for i in range(1, 6):
        state.po_registry.register_combo(f"PO_MANY_{i}", "00010", "+001", "001/001")

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    assert len(warnings) == 1
    _title, msg = warnings[0]
    assert "Phát hiện 5 dòng" in msg
    assert "và 2 dòng khác" in msg
    assert len(state.records) == 5

    state.po_registry.close()


def test_import_from_excel_auto_fill_po_and_duplicate_check(tk_root, tmp_path, monkeypatch):
    """Test importing Excel with empty POs triggers auto-fill before duplicate check."""
    excel_file = tmp_path / "empty_po.xlsx"
    _create_test_excel(
        excel_file,
        [
            {"item_code": "ITEM01", "item_name": "Part 1", "po": "", "po_detail": "", "po_sub": "", "box": "001/001"},
            {"item_code": "ITEM02", "item_name": "Part 2", "po": "", "po_detail": "", "po_sub": "", "box": "001/002"},
        ],
    )

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    # Newly generated POs should not be duplicates
    assert len(warnings) == 0
    assert len(state.records) == 2
    assert state.records[0].po.startswith("11")
    assert state.records[0].po_detail == FIXED_PO_DETAIL
    assert state.records[0].po_sub == FIXED_PO_SUB
    assert any("Đã tự động sinh PO cho 2 dòng" in log for log in mock_view.logs)

    state.po_registry.close()


def test_import_from_excel_headless_without_view(tk_root, tmp_path, monkeypatch):
    """Test import_from_excel works safely in headless mode where view is None."""
    excel_file = tmp_path / "headless.xlsx"
    _create_test_excel(
        excel_file,
        [
            {"item_code": "ITEM01", "item_name": "Part 1", "po": "PO_HEADLESS", "po_detail": "00010", "po_sub": "+001", "box": "001/001"},
        ],
    )

    state = AppState(tk_root)
    controller = AppController(state)
    controller.view = None  # Headless

    state.po_registry.register_combo("PO_HEADLESS", "00010", "+001", "001/001")

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    assert len(warnings) == 1
    assert len(state.records) == 1
    assert state.records[0].po == "PO_HEADLESS"

    state.po_registry.close()


def test_import_from_excel_empty_path_warning(tk_root, monkeypatch):
    """Test importing with empty path shows prompt warning."""
    state = AppState(tk_root)
    controller = AppController(state)

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set("")
    controller.import_from_excel()

    assert len(warnings) == 1
    assert "chưa chọn đường dẫn file excel" in warnings[0][1].lower()
    assert len(state.records) == 0

    state.po_registry.close()


def test_import_from_excel_empty_sheet_loads_empty_records(tk_root, tmp_path, monkeypatch):
    """Test importing an empty Excel file gracefully results in 0 records and no warning."""
    excel_file = tmp_path / "empty_sheet.xlsx"
    _create_test_excel(excel_file, [])

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    assert len(warnings) == 0
    assert len(state.records) == 0
    assert any("Đã import 0 dòng từ Excel" in log for log in mock_view.logs)

    state.po_registry.close()


def test_import_from_excel_invalid_revision_stops_before_duplicate_check(tk_root, tmp_path, monkeypatch):
    """Test importing with invalid revision shows error dialog and does not modify records."""
    excel_file = tmp_path / "invalid_rev.xlsx"
    _create_test_excel(
        excel_file,
        [
            {"item_code": "ITEM01", "item_name": "Part 1", "po": "PO_INV_01", "po_detail": "00010", "po_sub": "+001", "box": "001/001", "rev": "INVALID_REV"},
        ],
    )

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    errors: list[tuple] = []
    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showerror", lambda title, msg: errors.append((title, msg)))
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    assert len(errors) == 1
    assert "Không thể đọc dữ liệu từ file Excel" in errors[0][1]
    assert len(warnings) == 0
    assert len(state.records) == 0
    assert any("Lỗi import Excel" in log for log in mock_view.logs)

    state.po_registry.close()


def test_import_from_excel_custom_po_detail_sub_duplicates(tk_root, tmp_path, monkeypatch):
    """Test duplicate check correctly matches custom PO details (e.g. split 10010, return 11010)."""
    excel_file = tmp_path / "custom_split_return.xlsx"
    _create_test_excel(
        excel_file,
        [
            {"item_code": "ITEM01", "item_name": "Split Part", "po": "PO_SPLIT_01", "po_detail": "10010", "po_sub": "+001", "box": "001/002"},
            {"item_code": "ITEM02", "item_name": "Return Part", "po": "PO_RET_02", "po_detail": "11010", "po_sub": "+002", "box": "001/001"},
        ],
    )

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    state.po_registry.register_combo("PO_SPLIT_01", "10010", "+001", "001/002")
    state.po_registry.register_combo("PO_RET_02", "11010", "+002", "001/001")

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    assert len(warnings) == 1
    _title, msg = warnings[0]
    assert "Phát hiện 2 dòng" in msg
    assert "PO_SPLIT_01" in msg
    assert "PO_RET_02" in msg
    assert len(state.records) == 2

    state.po_registry.close()


def test_import_from_excel_intra_batch_duplicates_in_db(tk_root, tmp_path, monkeypatch):
    """Test importing Excel where multiple rows have the same registered combo."""
    excel_file = tmp_path / "batch_dup.xlsx"
    _create_test_excel(
        excel_file,
        [
            {"item_code": "ITEM01", "item_name": "Row 1", "po": "PO_SAME", "po_detail": "00010", "po_sub": "+001", "box": "001/001"},
            {"item_code": "ITEM02", "item_name": "Row 2", "po": "PO_SAME", "po_detail": "00010", "po_sub": "+001", "box": "001/001"},
        ],
    )

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    state.po_registry.register_combo("PO_SAME", "00010", "+001", "001/001")

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    assert len(warnings) == 1
    _title, msg = warnings[0]
    assert "Phát hiện 2 dòng" in msg
    assert len(state.records) == 2

    state.po_registry.close()


def test_import_from_excel_mixed_empty_po_and_duplicates(tk_root, tmp_path, monkeypatch):
    """Test importing Excel containing a mix of empty POs, duplicate POs, and clean POs."""
    excel_file = tmp_path / "mixed_po.xlsx"
    _create_test_excel(
        excel_file,
        [
            {"item_code": "ITEM01", "item_name": "Part 1 (empty PO)", "po": "", "po_detail": "", "po_sub": "", "box": "001/001"},
            {"item_code": "ITEM02", "item_name": "Part 2 (duplicate)", "po": "PO_EXIST_02", "po_detail": "00010", "po_sub": "+001", "box": "001/001"},
            {"item_code": "ITEM03", "item_name": "Part 3 (clean)", "po": "PO_FRESH_03", "po_detail": "00010", "po_sub": "+001", "box": "001/001"},
        ],
    )

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    state.po_registry.register_combo("PO_EXIST_02", "00010", "+001", "001/001")

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    # 1. Warning only flags PO_EXIST_02 (1 duplicate)
    assert len(warnings) == 1
    _title, msg = warnings[0]
    assert "Phát hiện 1 dòng" in msg
    assert "PO_EXIST_02" in msg
    assert "PO_FRESH_03" not in msg

    # 2. All 3 records loaded
    assert len(state.records) == 3
    assert state.records[0].po.startswith("11")  # Auto-generated
    assert state.records[1].po == "PO_EXIST_02"
    assert state.records[2].po == "PO_FRESH_03"

    # 3. Log captured auto-fill and duplicate warning
    assert any("Đã tự động sinh PO cho 1 dòng" in log for log in mock_view.logs)
    assert any("Cảnh báo: Phát hiện 1 dòng trùng mã EDI" in log for log in mock_view.logs)

    state.po_registry.close()


def test_import_from_excel_unnormalized_box_formats_matching_db(tk_root, tmp_path, monkeypatch):
    """Test importing Excel where box is unpadded (e.g. '1/3') matches standard '001/003' in DB."""
    excel_file = tmp_path / "unnorm_box.xlsx"
    _create_test_excel(
        excel_file,
        [
            {"item_code": "ITEM01", "item_name": "Part 1", "po": "PO_UNNORM", "po_detail": "00010", "po_sub": "+001", "box": "1/3"},
        ],
    )

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    # Pre-register normalized format '001/003'
    state.po_registry.register_combo("PO_UNNORM", "00010", "+001", "001/003")

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    # Box '1/3' is normalized to '001/003' by read_records/create_record, so duplicate is detected
    assert len(warnings) == 1
    _title, msg = warnings[0]
    assert "Phát hiện 1 dòng" in msg
    assert "PO_UNNORM" in msg
    assert "001/003" in msg
    assert len(state.records) == 1
    assert state.records[0].box == "001/003"

    state.po_registry.close()


def test_treeview_tag_configuration_and_duplicate_highlighting(tk_root, tmp_path):
    """Requirement R2: Treeview configures 'duplicate' tag with red styling and tags duplicate rows."""
    from tkinter import ttk
    from ui.components.data_tab import DataTabPanel
    from core.slip_printer_engine import create_record

    state = AppState(tk_root)
    controller = AppController(state)
    notebook = ttk.Notebook(tk_root)
    notebook.pack(fill="both", expand=True)

    data_tab = DataTabPanel(notebook, controller)
    notebook.add(data_tab, text="Data")
    tk_root.update_idletasks()

    # 1. Verify tag configure exists on Treeview
    tag_bg = data_tab.preview_tree.tag_configure("duplicate", "background")
    tag_fg = data_tab.preview_tree.tag_configure("duplicate", "foreground")
    assert str(tag_bg).lower() == "#fee2e2"
    assert str(tag_fg).lower() == "#991b1b"

    # 2. Register PO_DUP in database
    state.po_registry.register_combo("PO_DUP", "00010", "+001", "001/001")

    # 3. Add 3 records: one DB duplicate, one clean, one intra-batch duplicate pair
    state.records = [
        create_record(row_number=1, item_code="ITEM01", item_name="Part 1", carton_qty="10", total_qty="10", po="PO_DUP", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        create_record(row_number=2, item_code="ITEM02", item_name="Part 2", carton_qty="10", total_qty="10", po="PO_CLEAN", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        create_record(row_number=3, item_code="ITEM03", item_name="Part 3", carton_qty="10", total_qty="10", po="PO_BATCH_DUP", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        create_record(row_number=4, item_code="ITEM04", item_name="Part 4", carton_qty="10", total_qty="10", po="PO_BATCH_DUP", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
    ]

    data_tab.set_records(select_index=0)
    tk_root.update_idletasks()

    # Verify tags on Treeview rows
    assert data_tab.preview_tree.item("0", "tags") == ("duplicate",)
    assert not data_tab.preview_tree.item("1", "tags") or data_tab.preview_tree.item("1", "tags") in ((), "")
    assert data_tab.preview_tree.item("2", "tags") == ("duplicate",)
    assert data_tab.preview_tree.item("3", "tags") == ("duplicate",)

    state.po_registry.close()


def test_manual_add_record_clean_no_confirmation_dialog(tk_root, monkeypatch):
    """Requirement R3: Adding a clean record does not trigger confirmation dialog."""
    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    askyesno_called = []
    monkeypatch.setattr("tkinter.messagebox.askyesno", lambda title, msg: askyesno_called.append((title, msg)) or True)

    state.item_code_var.set("ITEM_CLEAN")
    state.item_name_var.set("Clean Part")
    state.carton_qty_var.set("10")
    state.box_var.set("001/001")
    state.rev_var.set("01")
    state.po_var.set("")  # Auto PO

    controller.add_record()

    assert len(askyesno_called) == 0
    assert len(state.records) == 1
    assert state.records[0].item_code == "ITEM_CLEAN"

    state.po_registry.close()


def test_manual_add_record_db_duplicate_prompt_user_cancels(tk_root, monkeypatch):
    """Requirement R3: When adding a record with duplicate in DB, user cancels (No), addition is aborted."""
    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    # Pre-register combo in DB
    state.po_registry.register_combo("1126081901", "00010", "+001", "001/001")

    prompts: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.askyesno", lambda title, msg: prompts.append((title, msg)) or False)

    state.item_code_var.set("ITEM_DUP")
    state.item_name_var.set("Dup Part")
    state.carton_qty_var.set("10")
    state.box_var.set("001/001")
    state.rev_var.set("01")
    state.po_var.set("1126081901")

    controller.add_record()

    assert len(prompts) == 1
    title, msg = prompts[0]
    assert title == APP_TITLE
    assert "CẢNH BÁO TRÙNG LẶP MÃ EDI" in msg
    assert "1126081901" in msg
    assert "trong cơ sở dữ liệu" in msg
    assert len(state.records) == 0
    assert any("Đã hủy thêm mới" in log for log in mock_view.logs)

    state.po_registry.close()


def test_manual_add_record_db_duplicate_prompt_user_confirms(tk_root, monkeypatch):
    """Requirement R3: When adding a record with duplicate in DB, user confirms (Yes), record is added."""
    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    state.po_registry.register_combo("1126081901", "00010", "+001", "001/001")

    prompts: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.askyesno", lambda title, msg: prompts.append((title, msg)) or True)

    state.item_code_var.set("ITEM_DUP")
    state.item_name_var.set("Dup Part")
    state.carton_qty_var.set("10")
    state.box_var.set("001/001")
    state.rev_var.set("01")
    state.po_var.set("1126081901")

    controller.add_record()

    assert len(prompts) == 1
    assert len(state.records) == 1
    assert state.records[0].po == "1126081901"

    state.po_registry.close()


def test_manual_add_consecutive_clicks_in_table_duplicate_prompt(tk_root, monkeypatch):
    """Requirement R3: Clicking 'Thêm mới' consecutively with the same PO/box triggers duplicate prompt."""
    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    prompts: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.askyesno", lambda title, msg: prompts.append((title, msg)) or False)

    state.item_code_var.set("ITEM_TEST")
    state.item_name_var.set("Test Part")
    state.carton_qty_var.set("10")
    state.box_var.set("001/001")
    state.rev_var.set("01")
    state.po_var.set("")

    # First add: clean auto PO
    controller.add_record()
    assert len(prompts) == 0
    assert len(state.records) == 1
    generated_po = state.records[0].po

    # Second add: po_var still holds generated_po, so clicking again detects duplicate in active table
    controller.add_record()
    assert len(prompts) == 1
    title, msg = prompts[0]
    assert "CẢNH BÁO TRÙNG LẶP MÃ EDI" in msg
    assert generated_po in msg
    assert "trong bảng hiện tại" in msg
    assert len(state.records) == 1  # User rejected, no second record added

    state.po_registry.close()


def test_import_warning_vietnamese_actionable_resolution(tk_root, tmp_path, monkeypatch):
    """Requirement R4: Import duplicate warning contains explicit guidance to delete red rows or edit box."""
    excel_file = tmp_path / "actionable_dup.xlsx"
    _create_test_excel(
        excel_file,
        [
            {"item_code": "ITEM01", "item_name": "Part 1", "po": "PO_WARN_01", "po_detail": "00010", "po_sub": "+001", "box": "001/001"},
        ],
    )

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    state.po_registry.register_combo("PO_WARN_01", "00010", "+001", "001/001")

    warnings: list[tuple] = []
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda title, msg: warnings.append((title, msg)))

    state.excel_var.set(str(excel_file))
    controller.import_from_excel()

    assert len(warnings) == 1
    title, msg = warnings[0]
    assert "⚠️ CẢNH BÁO TRÙNG LẶP MÃ EDI:" in msg
    assert "Các dòng bị trùng đã được bôi màu ĐỎ trên bảng dữ liệu" in msg
    assert "Vui lòng chọn dòng màu đỏ và nhấn 'Xóa dòng' hoặc đổi lại số Box" in msg

    state.po_registry.close()


