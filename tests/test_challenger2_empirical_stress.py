from __future__ import annotations

import re
from pathlib import Path
from tkinter import ttk
import openpyxl
import pytest

from core.po_registry import FIXED_PO_DETAIL, FIXED_PO_SUB, PORegistry
from core.slip_printer_engine import START_ROW, SlipRecord, create_record
from ui.app_controller import APP_TITLE, AppController
from ui.app_state import AppState
from ui.components.data_tab import DataTabPanel


class MockView:
    def __init__(self, state: AppState):
        self.app_state = state
        self.logs: list[str] = []
        self.records_set: list[SlipRecord] = []
        self.last_select_index: int | None = None
        self.generation_started = False
        self.preview_refreshed = False

    def append_log(self, message: str) -> None:
        self.logs.append(message)

    def set_records(self, records: list[SlipRecord], select_index: int | None = None, source: str | None = None) -> None:
        self.app_state.records = list(records)
        self.records_set = list(records)
        self.last_select_index = select_index
        if source:
            self.append_log(source)

    def refresh_preview_image(self) -> None:
        self.preview_refreshed = True

    def update_preview_display(self) -> None:
        pass

    def on_generation_start(self) -> None:
        self.generation_started = True

    def auto_commit_form(self) -> bool:
        return False


def _create_excel(file_path: Path, rows_data: list[dict[str, str]]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r_idx, data in enumerate(rows_data, start=START_ROW):
        ws.cell(row=r_idx, column=1, value=data.get("item_code", "ITEM001"))
        ws.cell(row=r_idx, column=2, value=data.get("item_name", "Test Item"))
        ws.cell(row=r_idx, column=3, value=data.get("total_qty", "100"))
        ws.cell(row=r_idx, column=4, value=data.get("carton_qty", "100"))
        ws.cell(row=r_idx, column=5, value=data.get("po", ""))
        ws.cell(row=r_idx, column=6, value=data.get("po_detail", "00010"))
        ws.cell(row=r_idx, column=7, value=data.get("po_sub", "+001"))
        ws.cell(row=r_idx, column=8, value=data.get("box", "001/001"))
        ws.cell(row=r_idx, column=9, value=data.get("rev", "01"))
        ws.cell(row=r_idx, column=10, value=data.get("lot", "LOT1234567"))
    wb.save(file_path)
    wb.close()
    return file_path


# =========================================================================
# 1. GRIDVIEW / TREEVIEW EDGE CASE DATASETS (REQUIREMENT R2)
# =========================================================================

def test_treeview_dataset_0_rows(tk_root):
    """Stress-test Treeview with 0 rows (empty dataset)."""
    state = AppState(tk_root)
    controller = AppController(state)
    notebook = ttk.Notebook(tk_root)
    notebook.pack(fill="both", expand=True)

    data_tab = DataTabPanel(notebook, controller)
    notebook.add(data_tab, text="Data")
    tk_root.update_idletasks()

    state.records = []
    data_tab.set_records(select_index=None)
    tk_root.update_idletasks()

    # Assertions on empty state
    assert len(data_tab.preview_tree.get_children()) == 0
    assert state.preview_index_map == []
    assert state.summary_var.get() == "Chưa có dữ liệu"
    assert "Thêm dữ liệu bằng form bên trái" in state.status_var.get()
    assert state.form_mode_var.get() == "Đang tạo dòng mới"
    assert state.preview_source_image is None
    assert data_tab.has_pending_form_changes() is False
    assert data_tab.auto_commit_form() is False

    state.po_registry.close()


def test_treeview_dataset_1_row_clean_vs_duplicate(tk_root):
    """Stress-test Treeview with 1 row: clean vs DB duplicate vs empty PO."""
    state = AppState(tk_root)
    controller = AppController(state)
    notebook = ttk.Notebook(tk_root)
    notebook.pack(fill="both", expand=True)

    data_tab = DataTabPanel(notebook, controller)
    notebook.add(data_tab, text="Data")
    tk_root.update_idletasks()

    # Case 1: 1 clean row
    state.records = [
        create_record(
            row_number=1, item_code="ITEM1", item_name="Item One", carton_qty="10",
            total_qty="10", po="1126000001", po_detail="00010", po_sub="+001",
            box="001/001", rev="01", lot="",
        ),
    ]
    data_tab.set_records(select_index=0)
    tk_root.update_idletasks()

    assert len(data_tab.preview_tree.get_children()) == 1
    assert data_tab.preview_tree.item("0", "tags") in ((), "")

    # Case 2: 1 duplicate row (in DB)
    state.po_registry.register_combo("1126000001", "00010", "+001", "001/001")
    data_tab.set_records(select_index=0)
    tk_root.update_idletasks()

    assert data_tab.preview_tree.item("0", "tags") == ("duplicate",)

    # Case 3: 1 row with empty PO
    state.records = [
        create_record(
            row_number=1, item_code="ITEM1", item_name="Item One", carton_qty="10",
            total_qty="10", po="", po_detail="00010", po_sub="+001",
            box="001/001", rev="01", lot="",
        ),
    ]
    data_tab.set_records(select_index=0)
    tk_root.update_idletasks()

    assert data_tab.preview_tree.item("0", "tags") in ((), "")

    state.po_registry.close()


def test_treeview_dataset_500_rows_and_pagination(tk_root):
    """Stress-test Treeview with 500 rows: performance, limit options, tagging integrity."""
    state = AppState(tk_root)
    controller = AppController(state)
    notebook = ttk.Notebook(tk_root)
    notebook.pack(fill="both", expand=True)

    data_tab = DataTabPanel(notebook, controller)
    notebook.add(data_tab, text="Data")
    tk_root.update_idletasks()

    # Pre-register 150 combos in DB
    for i in range(1, 151):
        state.po_registry.register_combo(f"PO_DB_{i:04d}", "00010", "+001", "001/001")

    records_500: list[SlipRecord] = []
    # 0..149: 150 rows matching DB duplicates
    for i in range(1, 151):
        records_500.append(
            create_record(
                row_number=i, item_code=f"ITEM_{i:04d}", item_name=f"DB Dup Item {i}",
                carton_qty="10", total_qty="10", po=f"PO_DB_{i:04d}", po_detail="00010",
                po_sub="+001", box="001/001", rev="01", lot="",
            )
        )
    # 150..249: 100 rows containing 50 intra-batch duplicate pairs (not in DB)
    for i in range(1, 51):
        po = f"PO_BATCH_{i:04d}"
        records_500.append(create_record(
            row_number=150 + (i * 2) - 1, item_code=f"ITEM_B_{i}", item_name=f"Batch Dup A {i}",
            carton_qty="10", total_qty="10", po=po, po_detail="00010", po_sub="+001",
            box="001/002", rev="01", lot="",
        ))
        records_500.append(create_record(
            row_number=150 + (i * 2), item_code=f"ITEM_B_{i}", item_name=f"Batch Dup B {i}",
            carton_qty="10", total_qty="10", po=po, po_detail="00010", po_sub="+001",
            box="001/002", rev="01", lot="",
        ))
    # 250..499: 250 completely clean unique records
    for i in range(1, 251):
        records_500.append(
            create_record(
                row_number=250 + i, item_code=f"ITEM_CLEAN_{i}", item_name=f"Clean Item {i}",
                carton_qty="10", total_qty="10", po=f"PO_CLEAN_{i:04d}", po_detail="00010",
                po_sub="+001", box="001/003", rev="01", lot="",
            )
        )

    assert len(records_500) == 500
    state.records = records_500

    # 1. Default preview_limit = 50
    state.preview_limit_var.set("50")
    data_tab.set_records(select_index=0)
    tk_root.update_idletasks()

    assert len(data_tab.preview_tree.get_children()) == 50
    assert len(state.preview_index_map) == 50
    assert state.summary_var.get() == "Có 500 dòng hợp lệ"
    assert "Đang hiển thị 50 dòng" in state.status_var.get()
    # All first 50 rows are from DB duplicates -> all must have ("duplicate",)
    for iid in data_tab.preview_tree.get_children():
        assert data_tab.preview_tree.item(iid, "tags") == ("duplicate",)

    # 2. Switch preview_limit = 500 (full view)
    state.preview_limit_var.set("500")
    data_tab.set_records(select_index=0)
    tk_root.update_idletasks()

    assert len(data_tab.preview_tree.get_children()) == 500
    assert len(state.preview_index_map) == 500

    # Check tagging across entire 500 dataset
    # Rows 0..149: DB duplicates
    for idx in range(150):
        assert data_tab.preview_tree.item(str(idx), "tags") == ("duplicate",), f"Row {idx} should be duplicate"

    # Rows 150..249: Intra-batch duplicates
    for idx in range(150, 250):
        assert data_tab.preview_tree.item(str(idx), "tags") == ("duplicate",), f"Row {idx} should be duplicate"

    # Rows 250..499: Clean rows
    for idx in range(250, 500):
        assert data_tab.preview_tree.item(str(idx), "tags") in ((), ""), f"Row {idx} should be clean"

    # 3. Select 500th row (index 499)
    data_tab._select_tree_row(499)
    tk_root.update_idletasks()
    assert state.selected_record_index == 499
    assert state.item_code_var.get() == "ITEM_CLEAN_250"
    assert state.form_mode_var.get() == "Đang sửa dòng 500"

    state.po_registry.close()


def test_treeview_dataset_100_percent_duplicates(tk_root):
    """Stress-test Treeview when 100% of rows are duplicates."""
    state = AppState(tk_root)
    controller = AppController(state)
    notebook = ttk.Notebook(tk_root)
    notebook.pack(fill="both", expand=True)

    data_tab = DataTabPanel(notebook, controller)
    notebook.add(data_tab, text="Data")
    tk_root.update_idletasks()

    # 100 rows all having the exact same PO and box
    state.records = [
        create_record(
            row_number=i, item_code=f"ITEM_{i}", item_name="Duplicate All",
            carton_qty="10", total_qty="10", po="1126999999", po_detail="00010",
            po_sub="+001", box="001/001", rev="01", lot="",
        )
        for i in range(1, 101)
    ]
    state.preview_limit_var.set("100")
    data_tab.set_records(select_index=0)
    tk_root.update_idletasks()

    children = data_tab.preview_tree.get_children()
    assert len(children) == 100
    for iid in children:
        assert data_tab.preview_tree.item(iid, "tags") == ("duplicate",)

    state.po_registry.close()


def test_treeview_mixed_duplicates_and_empty_po_combinations(tk_root):
    """Stress-test all combination cases: DB only, batch only, DB+batch, clean, empty PO, custom po_detail."""
    state = AppState(tk_root)
    controller = AppController(state)
    notebook = ttk.Notebook(tk_root)
    notebook.pack(fill="both", expand=True)

    data_tab = DataTabPanel(notebook, controller)
    notebook.add(data_tab, text="Data")
    tk_root.update_idletasks()

    # Pre-register combos in DB
    state.po_registry.register_combo("PO_DB_ONLY", "00010", "+001", "001/001")
    state.po_registry.register_combo("PO_DB_AND_BATCH", "00010", "+001", "001/001")
    state.po_registry.register_combo("PO_SPLIT_CUSTOM", "10010", "+002", "005/010")

    records = [
        # 0: DB only (single instance in batch) -> "duplicate"
        create_record(row_number=1, item_code="ITEM0", item_name="DB Only", carton_qty="10", total_qty="10", po="PO_DB_ONLY", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        # 1, 2: Batch only (not in DB, 2 instances) -> "duplicate"
        create_record(row_number=2, item_code="ITEM1", item_name="Batch Dup 1", carton_qty="10", total_qty="10", po="PO_BATCH_ONLY", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        create_record(row_number=3, item_code="ITEM2", item_name="Batch Dup 2", carton_qty="10", total_qty="10", po="PO_BATCH_ONLY", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        # 3, 4: DB and Batch (in DB and 2 instances) -> "duplicate"
        create_record(row_number=4, item_code="ITEM3", item_name="DB & Batch 1", carton_qty="10", total_qty="10", po="PO_DB_AND_BATCH", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        create_record(row_number=5, item_code="ITEM4", item_name="DB & Batch 2", carton_qty="10", total_qty="10", po="PO_DB_AND_BATCH", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        # 5: Completely clean -> ()
        create_record(row_number=6, item_code="ITEM5", item_name="Clean Unique", carton_qty="10", total_qty="10", po="PO_CLEAN_X", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        # 6, 7: Empty PO (multiple instances) -> () because empty PO is excluded from duplicate tagging
        create_record(row_number=7, item_code="ITEM6", item_name="Empty PO 1", carton_qty="10", total_qty="10", po="", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        create_record(row_number=8, item_code="ITEM7", item_name="Empty PO 2", carton_qty="10", total_qty="10", po="", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        # 8: Custom PO detail/sub in DB -> "duplicate"
        create_record(row_number=9, item_code="ITEM8", item_name="Custom Detail", carton_qty="10", total_qty="10", po="PO_SPLIT_CUSTOM", po_detail="10010", po_sub="+002", box="005/010", rev="01", lot=""),
    ]
    state.records = records
    data_tab.set_records(select_index=0)
    tk_root.update_idletasks()

    expected_tags = [
        ("duplicate",),  # 0
        ("duplicate",),  # 1
        ("duplicate",),  # 2
        ("duplicate",),  # 3
        ("duplicate",),  # 4
        (),              # 5
        (),              # 6
        (),              # 7
        ("duplicate",),  # 8
    ]

    for idx, expected in enumerate(expected_tags):
        actual = data_tab.preview_tree.item(str(idx), "tags")
        if not expected or expected == ():
            assert actual in ((), ""), f"Record at index {idx} ({records[idx].po}) expected untagged, got {actual}"
        else:
            assert actual == expected, f"Record at index {idx} ({records[idx].po}) expected tags {expected}, got {actual}"

    state.po_registry.close()


# =========================================================================
# 2. RAPID MANUAL ADD CLICKS & EDGE CASES (REQUIREMENT R3)
# =========================================================================

def test_rapid_manual_add_same_record_five_consecutive_clicks(tk_root, monkeypatch):
    """Stress-test 5 consecutive clicks on '➕ Thêm mới' with same form data."""
    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    prompts: list[tuple[str, str]] = []
    responses = [False, False, True, True]  # For clicks 2, 3, 4, 5

    def mock_askyesno(title, msg):
        prompts.append((title, msg))
        return responses.pop(0)

    monkeypatch.setattr("tkinter.messagebox.askyesno", mock_askyesno)

    state.item_code_var.set("RAPID_ITEM")
    state.item_name_var.set("Rapid Test Item")
    state.carton_qty_var.set("15")
    state.box_var.set("001/001")
    state.rev_var.set("01")
    state.po_var.set("")  # Auto-generated on first click

    # 1st click: clean auto PO, 0 prompts, records count = 1
    controller.add_record()
    assert len(prompts) == 0
    assert len(state.records) == 1
    auto_po = state.records[0].po
    assert auto_po.startswith("11")
    assert state.po_var.get() == auto_po

    # 2nd click: user rejects (False) -> prompt shown, records count remains 1
    controller.add_record()
    assert len(prompts) == 1
    assert "CẢNH BÁO TRÙNG LẶP MÃ EDI" in prompts[0][1]
    assert auto_po in prompts[0][1]
    assert "trong bảng hiện tại" in prompts[0][1]
    assert len(state.records) == 1

    # 3rd click: user rejects again (False) -> prompt shown, records count remains 1
    controller.add_record()
    assert len(prompts) == 2
    assert len(state.records) == 1

    # 4th click: user accepts (True) -> prompt shown, records count becomes 2
    controller.add_record()
    assert len(prompts) == 3
    assert len(state.records) == 2
    assert state.records[1].po == auto_po
    assert state.records[1].row_number == state.records[0].row_number + 1

    # 5th click: user accepts (True) -> prompt shown, records count becomes 3
    controller.add_record()
    assert len(prompts) == 4
    assert len(state.records) == 3
    assert state.records[2].po == auto_po
    assert state.records[2].row_number == state.records[0].row_number + 2

    # Verify that in Treeview, all 3 rows get tagged as duplicate
    notebook = ttk.Notebook(tk_root)
    data_tab = DataTabPanel(notebook, controller)
    notebook.add(data_tab, text="Data")
    data_tab.set_records(select_index=0)
    tk_root.update_idletasks()

    assert len(data_tab.preview_tree.get_children()) == 3
    for iid in ("0", "1", "2"):
        assert data_tab.preview_tree.item(iid, "tags") == ("duplicate",)

    state.po_registry.close()


def test_manual_add_box_range_expansion_with_partial_db_duplicates(tk_root, monkeypatch):
    """Stress-test manual addition with box range (e.g. 001-004) where some boxes exist in DB."""
    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    # Pre-register box 002/004 and 004/004 in DB for PO 1126123456
    po = "1126123456"
    state.po_registry.register_combo(po, "00010", "+001", "002/004")
    state.po_registry.register_combo(po, "00010", "+001", "004/004")

    prompts: list[tuple[str, str]] = []
    user_confirm = False

    def mock_askyesno(title, msg):
        prompts.append((title, msg))
        return user_confirm

    monkeypatch.setattr("tkinter.messagebox.askyesno", mock_askyesno)

    state.item_code_var.set("MULTI_BOX_ITEM")
    state.item_name_var.set("Multi Box Part")
    state.carton_qty_var.set("20")
    state.box_var.set("4")  # 4 boxes
    state.rev_var.set("01")
    state.po_var.set(po)

    # 1. User cancels (user_confirm = False)
    user_confirm = False
    controller.add_record()

    assert len(prompts) == 1
    assert "2 dòng tem vừa tạo đã tồn tại" in prompts[0][1]
    assert "Box: 002/004 (trong cơ sở dữ liệu)" in prompts[0][1]
    assert "Box: 004/004 (trong cơ sở dữ liệu)" in prompts[0][1]
    assert len(state.records) == 0

    # 2. User confirms (user_confirm = True)
    user_confirm = True
    controller.add_record()

    assert len(prompts) == 2
    assert len(state.records) == 4
    assert [r.box for r in state.records] == ["001/004", "002/004", "003/004", "004/004"]

    # Verify UI tags
    notebook = ttk.Notebook(tk_root)
    data_tab = DataTabPanel(notebook, controller)
    notebook.add(data_tab, text="Data")
    data_tab.set_records(select_index=0)
    tk_root.update_idletasks()

    assert data_tab.preview_tree.item("0", "tags") in ((), "")        # 001/004 clean
    assert data_tab.preview_tree.item("1", "tags") == ("duplicate",) # 002/004 dup
    assert data_tab.preview_tree.item("2", "tags") in ((), "")        # 003/004 clean
    assert data_tab.preview_tree.item("3", "tags") == ("duplicate",) # 004/004 dup

    state.po_registry.close()


def test_manual_add_validation_errors_fail_safely(tk_root, monkeypatch):
    """Stress-test manual addition with missing fields and invalid rev."""
    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    errors: list[tuple[str, str]] = []
    monkeypatch.setattr("tkinter.messagebox.showerror", lambda title, msg: errors.append((title, msg)))

    # Test cases: (item_code, item_name, carton_qty, box, rev, expected_error_substr)
    invalid_cases = [
        ("", "Item Name", "10", "001/001", "01", "Bạn chưa nhập Mã hàng"),
        ("ITEM", "", "10", "001/001", "01", "Bạn chưa nhập Tên hàng"),
        ("ITEM", "Item Name", "", "001/001", "01", "Bạn chưa nhập Số lượng thùng"),
        ("ITEM", "Item Name", "10", "", "01", "Bạn chưa nhập Số box"),
        ("ITEM", "Item Name", "10", "001/001", "ABC", "Rev phải có 2 chữ số"),
        ("ITEM", "Item Name", "10", "001/001", "0", "Rev phải có 2 chữ số"),
        ("ITEM", "Item Name", "10", "001/001", "000", "Rev phải có 2 chữ số"),
    ]

    for item_code, item_name, carton_qty, box, rev, err_sub in invalid_cases:
        errors.clear()
        state.item_code_var.set(item_code)
        state.item_name_var.set(item_name)
        state.carton_qty_var.set(carton_qty)
        state.box_var.set(box)
        state.rev_var.set(rev)

        controller.add_record()

        assert len(errors) == 1, f"Failed for case ({item_code}, {item_name}, {carton_qty}, {box}, {rev})"
        assert err_sub in errors[0][1]
        assert "👉 Hướng dẫn: Vui lòng kiểm tra và nhập đầy đủ các trường bắt buộc" in errors[0][1]
        assert len(state.records) == 0

    state.po_registry.close()


# =========================================================================
# 3. DIALOG BRANCHING VERIFICATION (REQUIREMENT R3)
# =========================================================================

def test_dialog_branching_rejection_leaves_state_strictly_intact(tk_root, monkeypatch):
    """Verify that when askyesno returns False, app_state.records is completely unmodified."""
    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    # Initial 2 records
    state.records = [
        create_record(row_number=1, item_code="ITEM_INIT_1", item_name="Initial Part 1", carton_qty="10", total_qty="10", po="PO_INIT_01", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
        create_record(row_number=2, item_code="ITEM_INIT_2", item_name="Initial Part 2", carton_qty="10", total_qty="10", po="PO_INIT_02", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
    ]
    snapshot_before = list(state.records)

    # Pre-register duplicate in DB
    state.po_registry.register_combo("PO_NEW_DUP", "00010", "+001", "001/001")

    prompts = []
    monkeypatch.setattr("tkinter.messagebox.askyesno", lambda title, msg: prompts.append((title, msg)) or False)

    state.item_code_var.set("NEW_DUP_ITEM")
    state.item_name_var.set("New Duplicate Item")
    state.carton_qty_var.set("50")
    state.box_var.set("001/001")
    state.rev_var.set("01")
    state.po_var.set("PO_NEW_DUP")

    controller.add_record()

    assert len(prompts) == 1
    assert len(state.records) == 2
    assert state.records == snapshot_before
    assert any("Đã hủy thêm mới do phát hiện trùng mã EDI" in log for log in mock_view.logs)

    state.po_registry.close()


def test_dialog_branching_acceptance_extends_state_and_updates_view(tk_root, monkeypatch):
    """Verify that when askyesno returns True, record is appended and view receives updated records."""
    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    state.records = [
        create_record(row_number=1, item_code="ITEM_INIT_1", item_name="Initial Part 1", carton_qty="10", total_qty="10", po="PO_INIT_01", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot=""),
    ]

    state.po_registry.register_combo("PO_NEW_DUP", "00010", "+001", "001/001")

    prompts = []
    monkeypatch.setattr("tkinter.messagebox.askyesno", lambda title, msg: prompts.append((title, msg)) or True)

    state.item_code_var.set("NEW_DUP_ITEM")
    state.item_name_var.set("New Duplicate Item")
    state.carton_qty_var.set("50")
    state.box_var.set("001/001")
    state.rev_var.set("01")
    state.po_var.set("PO_NEW_DUP")

    controller.add_record()

    assert len(prompts) == 1
    assert len(state.records) == 2
    assert state.records[1].po == "PO_NEW_DUP"
    assert state.records[1].item_code == "NEW_DUP_ITEM"
    assert len(mock_view.records_set) == 2
    assert mock_view.last_select_index == 1
    assert any("Đã thêm dòng 2: NEW_DUP_ITEM" in log for log in mock_view.logs)

    state.po_registry.close()


# =========================================================================
# 4. VIETNAMESE DIACRITICS & ACTIONABLE GUIDANCE (REQUIREMENT R4)
# =========================================================================

def _has_vietnamese_diacritics(text: str) -> bool:
    viet_pattern = re.compile(r"[àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡùúụủũưừứựửữỳýỵỷỹđĐ]")
    return bool(viet_pattern.search(text))


def _has_corrupt_mojibake(text: str) -> bool:
    mojibake_indicators = ["Ã¡", "Ã©", "Ã", "â€", "\ufffd"]
    return any(indicator in text for indicator in mojibake_indicators)


def _has_actionable_guidance(text: str) -> bool:
    actionable_markers = ["Hướng dẫn", "Lưu ý", "👉", "Vui lòng", "Bạn có", "Chọn"]
    return any(marker in text for marker in actionable_markers)


def test_vietnamese_diacritics_and_actionable_guidance_in_all_controller_dialogs(tk_root, monkeypatch):
    """Empirically trigger every messagebox dialog in AppController and verify Vietnamese text and guidance."""
    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    captured_dialogs: list[dict[str, str]] = []

    def mock_showwarning(title, msg):
        captured_dialogs.append({"type": "showwarning", "title": str(title), "msg": str(msg)})

    def mock_showerror(title, msg):
        captured_dialogs.append({"type": "showerror", "title": str(title), "msg": str(msg)})

    def mock_showinfo(title, msg):
        captured_dialogs.append({"type": "showinfo", "title": str(title), "msg": str(msg)})

    def mock_askyesno(title, msg):
        captured_dialogs.append({"type": "askyesno", "title": str(title), "msg": str(msg)})
        return False

    monkeypatch.setattr("tkinter.messagebox.showwarning", mock_showwarning)
    monkeypatch.setattr("tkinter.messagebox.showerror", mock_showerror)
    monkeypatch.setattr("tkinter.messagebox.showinfo", mock_showinfo)
    monkeypatch.setattr("tkinter.messagebox.askyesno", mock_askyesno)

    # 1. warn_lot_field_locked
    controller.warn_lot_field_locked()

    # 2. add_record invalid input
    state.item_code_var.set("")
    controller.add_record()

    # 3. update_selected_record on empty table
    state.records = []
    controller.update_selected_record()

    # 4. clear_all_records confirmation
    state.records = [create_record(row_number=1, item_code="ITEM", item_name="Item", carton_qty="10", total_qty="10", po="1126000001", po_detail="00010", po_sub="+001", box="001/001", rev="01", lot="")]
    controller.clear_all_records()

    # 5. import_from_excel empty path
    state.excel_var.set("")
    controller.import_from_excel()

    # 6. import_from_excel non-existent file
    state.excel_var.set("non_existent_file_path.xlsx")
    controller.import_from_excel()

    # 7. nudge_layout invalid coords
    state.layout_choice_var.set("item_code")
    state.x_var.set("invalid_number")
    controller.nudge_layout(1.0, 0.0)

    # 8. resize_layout invalid coords
    state.width_var.set("5.0")
    state.height_var.set("not_a_float")
    controller.resize_layout(1.0, 1.0)

    # 9. save_layout_config_to_disk
    controller.save_layout_config_to_disk()

    # 10. start_generation empty records
    state.records = []
    controller.start_generation()

    # 11. open_generated_pdf when none generated
    state.generated_output_path = None
    controller.open_generated_pdf()

    # 12. open_build_script when non-existent
    state.bundle_dir = Path("non_existent_folder")
    controller.open_build_script()

    # Verify each dialog captured
    assert len(captured_dialogs) >= 12

    for dlg in captured_dialogs:
        msg = dlg["msg"]
        title = dlg["title"]

        # Check title
        assert title in (APP_TITLE, "In Phiếu Hiện Vật"), f"Unexpected title: {title}"

        # Check Vietnamese diacritics
        assert _has_vietnamese_diacritics(msg), f"Dialog message missing Vietnamese diacritics: {msg}"

        # Check no mojibake
        assert not _has_corrupt_mojibake(msg), f"Dialog message contains mojibake: {msg}"

        # Check actionable guidance
        assert _has_actionable_guidance(msg), f"Dialog message missing actionable guidance marker: {msg}"

    state.po_registry.close()


def test_vietnamese_diacritics_in_qr_scan_dialog(tk_root, monkeypatch):
    """Empirically test QRScanDialog error handling and Vietnamese strings."""
    from ui.components.qr_scan_dialog import QRScanDialog

    state = AppState(tk_root)
    controller = AppController(state)
    mock_view = MockView(state)
    controller.set_view(mock_view)

    dialog = QRScanDialog(tk_root, controller)
    tk_root.update_idletasks()

    dialog_errors: list[dict[str, str]] = []
    dialog_warnings: list[dict[str, str]] = []

    monkeypatch.setattr("tkinter.messagebox.showerror", lambda t, m: dialog_errors.append({"title": t, "msg": m}))
    monkeypatch.setattr("tkinter.messagebox.showwarning", lambda t, m: dialog_warnings.append({"title": t, "msg": m}))

    # 1. Process empty QR code
    dialog.scan_input_var.set("")
    dialog.process_scanned_code()
    assert len(dialog_warnings) == 1
    assert _has_vietnamese_diacritics(dialog_warnings[0]["msg"])
    assert _has_actionable_guidance(dialog_warnings[0]["msg"])

    # 2. Process invalid format QR code
    dialog.scan_input_var.set("INVALID_SHORT_CODE")
    dialog.process_scanned_code()
    assert len(dialog_errors) == 1
    assert _has_vietnamese_diacritics(dialog_errors[0]["msg"])
    assert _has_actionable_guidance(dialog_errors[0]["msg"])

    # 3. Confirm and add records with empty required fields
    dialog.item_code_var.set("")
    dialog.confirm_and_add_records()
    assert len(dialog_errors) == 2
    assert _has_vietnamese_diacritics(dialog_errors[1]["msg"])
    assert _has_actionable_guidance(dialog_errors[1]["msg"])

    dialog.destroy()
    state.po_registry.close()
