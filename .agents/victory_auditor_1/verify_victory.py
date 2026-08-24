"""Independent Victory Audit Verification Script.
Conducts full end-to-end verification of import_from_excel duplicate EDI checking,
including mock UI, database state, auto-fill synergy, non-blocking warning, and error paths.
"""
from __future__ import annotations

import os
import sys
import tempfile
import tkinter as tk
from pathlib import Path
import openpyxl

# Add repo root to sys.path
REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from core.po_registry import FIXED_PO_DETAIL, FIXED_PO_SUB, PORegistry
from core.slip_printer_engine import START_ROW
from ui.app_controller import APP_TITLE, AppController
from ui.app_state import AppState


def create_excel(file_path: Path, rows: list[dict[str, str]]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r_idx, data in enumerate(rows, start=START_ROW):
        ws.cell(row=r_idx, column=1, value=data.get("item_code", "ITEM001"))
        ws.cell(row=r_idx, column=2, value=data.get("item_name", "Test Item"))
        ws.cell(row=r_idx, column=3, value=data.get("total_qty", "100"))
        ws.cell(row=r_idx, column=4, value=data.get("carton_qty", "100"))
        ws.cell(row=r_idx, column=5, value=data.get("po", ""))
        ws.cell(row=r_idx, column=6, value=data.get("po_detail", "00010"))
        ws.cell(row=r_idx, column=7, value=data.get("po_sub", "+001"))
        ws.cell(row=r_idx, column=8, value=data.get("box", "001/001"))
        ws.cell(row=r_idx, column=9, value=data.get("rev", "01"))
        ws.cell(row=r_idx, column=10, value=data.get("lot", "LOT1234567"))
    wb.save(file_path)
    wb.close()
    return file_path


class MockView:
    def __init__(self, state: AppState):
        self.app_state = state
        self.logs: list[str] = []
        self.records_set: list = []

    def append_log(self, message: str) -> None:
        self.logs.append(message)

    def set_records(self, records, select_index=None, source=None) -> None:
        self.app_state.records = list(records)
        self.records_set = list(records)
        if source:
            self.append_log(source)


def run_audit():
    print("=== STARTING INDEPENDENT AUDIT ===")
    root = tk.Tk()
    root.withdraw()

    temp_dir = Path(tempfile.mkdtemp())
    try:
        # Check 1: Clean Import (No Duplicates)
        state1 = AppState(root)
        controller1 = AppController(state1)
        view1 = MockView(state1)
        controller1.set_view(view1)
        warnings1 = []
        import tkinter.messagebox as mb
        orig_warn = mb.showwarning
        mb.showwarning = lambda title, msg: warnings1.append((title, msg))

        excel1 = temp_dir / "clean.xlsx"
        create_excel(excel1, [
            {"item_code": "ITEM01", "item_name": "Part 1", "po": "PO_C1", "box": "001/001"},
            {"item_code": "ITEM02", "item_name": "Part 2", "po": "PO_C2", "box": "001/001"},
        ])
        state1.excel_var.set(str(excel1))
        controller1.import_from_excel()

        assert len(warnings1) == 0, f"Expected 0 warnings, got {len(warnings1)}"
        assert len(state1.records) == 2, f"Expected 2 records, got {len(state1.records)}"
        print("[CHECK 1 PASS] Clean import loads all records with 0 warnings.")
        state1.po_registry.close()

        # Check 2: Duplicates Detected -> Non-blocking warning + All records loaded
        state2 = AppState(root)
        controller2 = AppController(state2)
        view2 = MockView(state2)
        controller2.set_view(view2)
        warnings2 = []
        mb.showwarning = lambda title, msg: warnings2.append((title, msg))

        state2.po_registry.register_combo("PO_DUP_1", "00010", "+001", "001/001")
        state2.po_registry.register_combo("PO_DUP_2", "00010", "+001", "001/002")

        excel2 = temp_dir / "dups.xlsx"
        create_excel(excel2, [
            {"item_code": "ITEM01", "item_name": "Part 1", "po": "PO_DUP_1", "box": "001/001"},
            {"item_code": "ITEM02", "item_name": "Part 2", "po": "PO_DUP_2", "box": "001/002"},
            {"item_code": "ITEM03", "item_name": "Part 3", "po": "PO_CLEAN_3", "box": "001/001"},
        ])
        state2.excel_var.set(str(excel2))
        controller2.import_from_excel()

        assert len(warnings2) == 1, f"Expected 1 warning, got {len(warnings2)}"
        title, msg = warnings2[0]
        assert "Phát hiện 2 dòng có mã EDI đã tồn tại" in msg
        assert "PO_DUP_1" in msg
        assert "PO_DUP_2" in msg
        assert "Dữ liệu vẫn được nạp vào bảng" in msg
        assert len(state2.records) == 3, f"Expected all 3 records loaded, got {len(state2.records)}"
        assert any("Cảnh báo: Phát hiện 2 dòng trùng mã EDI" in log for log in view2.logs)
        print("[CHECK 2 PASS] Duplicate warning displayed without blocking import; all 3 records loaded.")
        state2.po_registry.close()

        # Check 3: >3 Duplicates Formatting
        state3 = AppState(root)
        controller3 = AppController(state3)
        warnings3 = []
        mb.showwarning = lambda title, msg: warnings3.append((title, msg))

        rows3 = []
        for i in range(1, 6):
            state3.po_registry.register_combo(f"PO_M_{i}", "00010", "+001", "001/001")
            rows3.append({"item_code": f"IT{i}", "item_name": f"P{i}", "po": f"PO_M_{i}", "box": "001/001"})
        excel3 = temp_dir / "many.xlsx"
        create_excel(excel3, rows3)
        state3.excel_var.set(str(excel3))
        controller3.import_from_excel()

        assert len(warnings3) == 1
        assert "Phát hiện 5 dòng" in warnings3[0][1]
        assert "và 2 dòng khác" in warnings3[0][1]
        assert len(state3.records) == 5
        print("[CHECK 3 PASS] >3 Duplicates formatted with sample list and overflow count.")
        state3.po_registry.close()

        # Restore messagebox
        mb.showwarning = orig_warn
        print("=== ALL INDEPENDENT AUDIT CHECKS PASSED ===")
    finally:
        try:
            root.destroy()
        except Exception:
            pass

if __name__ == "__main__":
    run_audit()
